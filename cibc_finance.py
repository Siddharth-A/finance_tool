#!/usr/bin/env python3

# External libraries
import sys
import os
import csv
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# Internal libraries
import categories

# Global variables
csv_visa = 'cibc-visa.csv'
csv_chq = 'cibc-chq.csv'
csv_sav = 'cibc-sav.csv'
csv_roger = 'rogers.csv'
output_file = 'cibc_master.xlsx'
output_sheet_title = 'banking'
orange = PatternFill(start_color='ffcaa1',end_color='ffcaa1',fill_type='solid')
purple = PatternFill(start_color='dbb5ff',end_color='dbb5ff',fill_type='solid')
red1   = PatternFill(start_color='ffa1a1',end_color='ffa1a1',fill_type='solid')
red2   = PatternFill(start_color='e06565',end_color='e06565',fill_type='solid')
red3   = PatternFill(start_color='ff2e2e',end_color='ff2e2e',fill_type='solid')
yellow = PatternFill(start_color='fffa99',end_color='fffa99',fill_type='solid')
green1 = PatternFill(start_color='c1ffab',end_color='c1ffab',fill_type='solid')
green2 = PatternFill(start_color='00ff80',end_color='00ff80',fill_type='solid')
green3 = PatternFill(start_color='34fa4f',end_color='34fa4f',fill_type='solid')
blue   = PatternFill(start_color='abf2ff',end_color='abf2ff',fill_type='solid')
white  = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')
black  = PatternFill(start_color='000000',end_color='000000',fill_type='solid')

# converts the downloaded csv files (cibc-chq, cibc-sav & cibc-visa) to xlsx format (one single file)
# and writes VISA/SAV/CHQ to column G depending on the input file the row is copied from
def create_master_xlsx(output_file):
	print ("\n1) Converting CSV files to XLSX and appending transacation type\n")
	wb = Workbook()
	ws = wb.active

	with open(csv_visa, 'r') as f:
	    for row in csv.reader(f):
	        ws.append(row)
	last_visa_row = ws.max_row
	i = 1
	while i <= last_visa_row:
		ws.cell(row=i, column=7).value = 'VISA'
		i += 1
	print ("	{} converted and transaction type appended (to column G)".format(csv_visa))

	with open(csv_chq, 'r') as f:
	    for row in csv.reader(f):
	        ws.append(row)
	last_chq_row = ws.max_row
	i = last_visa_row + 1
	while i <= last_chq_row:
		ws.cell(row=i, column=7).value = 'CHQ'
		i += 1
	print ("	{} converted and transaction type appended (to column G)".format(csv_chq))

	with open(csv_sav, 'r') as f:
	    for row in csv.reader(f):
	        ws.append(row)
	last_sav_row = ws.max_row
	i = last_chq_row + 1
	while i <= last_sav_row:
		ws.cell(row=i, column=7).value = 'SAV'
		i += 1
	print ("	{} converted and transaction type appended (to column G)".format(csv_sav))

	with open(csv_roger, 'r') as f:
	    for row in csv.reader(f):
	        ws.append(row)
	last_roger_row = ws.max_row
	i = last_sav_row + 1
	while i <= last_roger_row:
		ws.cell(row=i, column=7).value = 'ROGER'
		i += 1
	print ("	{} converted and transaction type appended (to column G)".format(csv_roger))

	ws.title = output_sheet_title
	wb.save(output_file)

# fills in cell color depending on it's value and the list it belongs to
def cell_color(input_file, worksheet):
	print ("\n2) Color coding each cell as per classification\n")
	no_match = 0
	no_match_list = []
	wb = load_workbook(input_file)
	ws = wb[worksheet]
	colB = ws['B']
	for cell_des in colB:
		cell = cell_des.value
		if any(word in cell for word in categories.grocery):
			cell_des.fill = red3
			continue

		elif any(word in cell for word in categories.resteraunts):
			cell_des.fill = red3
			continue

		elif any(word in cell for word in categories.entertain):
			cell_des.fill = red3
			continue

		elif any(word in cell for word in categories.transport):
			cell_des.fill = orange
			continue

		elif any(word in cell for word in categories.personal):
			cell_des.fill = orange
			continue

		elif any(word in cell for word in categories.retail):
			cell_des.fill = orange
			continue

		elif any(word in cell for word in categories.credit):
			cell_des.fill = green2
			continue

		elif any(word in cell for word in categories.pay):
			cell_des.fill = green2
			continue

		elif any(word in cell for word in categories.savings):
			cell_des.fill = green2
			continue

		else:
			no_match_list.append(cell_des)
			no_match+=1
			continue

	print ("	Number of unmatched entries: ", no_match)
	#print (*no_match_list, sep = "\n")
	wb.save(input_file)

# Performs basic formatting on the output XLSX file
def format_sheet(input_file, worksheet):
	wb = load_workbook(input_file)
	ws = wb[worksheet]
	print ("\n4) Formating worksheet: {}\n".format(worksheet))
	ws.column_dimensions['A'].width = 10
	ws.column_dimensions['B'].width = 42

	colC = ws['C']
	colD = ws['D']
	colG = ws['G']
	for cell_des in colC:
		cell_des.alignment = Alignment(horizontal="center", vertical="center")
	for cell_des in colD:
		cell_des.alignment = Alignment(horizontal="center", vertical="center")
	for cell_des in colG:
		cell_des.alignment = Alignment(horizontal="center", vertical="center")
	wb.save(input_file)


def duplicate_entry(input_file, worksheet):
	print ("\n5) Highlighting internet transfers from CHQ to SAV\n")
	wb = load_workbook(input_file)
	ws = wb[worksheet]
	i = 1
	j = 2
	last_row = ws.max_row
	while i <= last_row:
		if ws.cell(row=i, column=2).value == ws.cell(row=j, column=2).value:
			ws.cell(row=i, column=2).fill = black
			ws.cell(row=j, column=2).fill = black
		i+=1
		j+=1
	wb.save(input_file)

def main():
	create_master_xlsx(output_file)
	cell_color(output_file, output_sheet_title)
	format_sheet(output_file, output_sheet_title)
	duplicate_entry(output_file, output_sheet_title)

if __name__== "__main__":
  main()




